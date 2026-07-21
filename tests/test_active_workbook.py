import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

from targetare_contacts import create_app
from targetare_contacts.db import get_db
from targetare_contacts.targetare import InterrogationResult


def workbook_stream():
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(["Denumire", "CUI", "Adresa"])
    worksheet.append(["ACME SRL", "RO12345678", "Bucuresti"])
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream


def make_config(tmp_path, database_name="db.sqlite3"):
    return {
        "TESTING": True,
        "SECRET_KEY": "test",
        "TARGETARE_API_KEY": "secret",
        "DATABASE": str(tmp_path / database_name),
        "ACTIVE_WORKBOOK": str(tmp_path / "active.xlsx"),
        "ACTIVE_WORKBOOK_BACKUP": str(tmp_path / "active.backup.xlsx"),
        "ACTIVE_WORKBOOK_META": str(tmp_path / "active.json"),
    }


def upload_original(client, follow_redirects=True):
    return client.post(
        "/upload",
        data={"file": (workbook_stream(), "lista-firme.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=follow_redirects,
    )


def install_fake_client(monkeypatch):
    class FakeClient:
        def __init__(self, **_kwargs):
            pass

        def interrogate(self, tax_id):
            assert tax_id == "12345678"
            return InterrogationResult(
                status="success",
                emails={"primaryEmail": "office@example.ro", "websiteEmails": []},
                phones={"primaryPhone": "+40700000000", "verifiedPhones": []},
                remaining_requests=77,
            )

    monkeypatch.setattr("targetare_contacts.TargetareClient", FakeClient)


def test_refresh_keeps_visible_active_filename(tmp_path):
    app = create_app(make_config(tmp_path))
    client = app.test_client()
    response = upload_original(client)
    assert b"lista-firme.xlsx" in response.data

    refreshed = client.get("/")
    assert b"Fi\xc8\x99ier activ" in refreshed.data
    assert b"lista-firme.xlsx" in refreshed.data
    assert b"R\xc4\x83m\xc3\xa2ne activ" in refreshed.data


def test_interrogation_is_blocked_without_active_workbook(tmp_path, monkeypatch):
    app = create_app(make_config(tmp_path))
    called = False

    class ShouldNotRun:
        def __init__(self, **_kwargs):
            nonlocal called
            called = True

    monkeypatch.setattr("targetare_contacts.TargetareClient", ShouldNotRun)
    with app.app_context():
        db = get_db()
        db.execute(
            "INSERT INTO companies (company_name, tax_id, source_row) VALUES (?, ?, ?)",
            ("ACME SRL", "12345678", 2),
        )
        db.commit()
        company_id = db.execute("SELECT id FROM companies").fetchone()["id"]

    response = app.test_client().post(
        f"/companies/{company_id}/interrogate", follow_redirects=True
    )
    assert b"Interogarea a fost blocat" in response.data
    assert called is False


def test_reuploading_stale_file_preserves_existing_results(tmp_path, monkeypatch):
    app = create_app(make_config(tmp_path))
    client = app.test_client()
    upload_original(client)
    install_fake_client(monkeypatch)

    with app.app_context():
        company_id = get_db().execute("SELECT id FROM companies").fetchone()["id"]
    client.post(f"/companies/{company_id}/interrogate", follow_redirects=True)

    response = upload_original(client)
    assert b"office@example.ro" in response.data
    assert b"p\xc4\x83strate 1 rezultate" in response.data

    workbook = load_workbook(app.config["ACTIVE_WORKBOOK"], data_only=True)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1]}
    assert worksheet.cell(2, headers["Emailuri Targetare"]).value == "office@example.ro"
    assert worksheet.cell(2, headers["Telefoane Targetare"]).value == "+40700000000"
    assert worksheet.cell(2, headers["Status interogare"]).value == "Interogat"
    workbook.close()
    assert Path(app.config["ACTIVE_WORKBOOK_BACKUP"]).is_file()


def test_restart_restores_active_workbook_without_upload(tmp_path, monkeypatch):
    first_app = create_app(make_config(tmp_path, "first.sqlite3"))
    client = first_app.test_client()
    upload_original(client)
    install_fake_client(monkeypatch)

    with first_app.app_context():
        company_id = get_db().execute("SELECT id FROM companies").fetchone()["id"]
    client.post(f"/companies/{company_id}/interrogate", follow_redirects=True)

    second_app = create_app(make_config(tmp_path, "second.sqlite3"))
    response = second_app.test_client().get("/")
    assert b"lista-firme.xlsx" in response.data
    assert b"office@example.ro" in response.data
    assert b"+40700000000" in response.data
    assert b"Reinterogheaz" in response.data


def test_download_keeps_original_filename(tmp_path):
    app = create_app(make_config(tmp_path))
    client = app.test_client()
    upload_original(client)
    response = client.get("/download")
    assert response.status_code == 200
    assert "lista-firme.xlsx" in response.headers["Content-Disposition"]
