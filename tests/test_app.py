import io

from openpyxl import Workbook, load_workbook

from targetare_contacts import create_app
from targetare_contacts.db import get_db
from targetare_contacts.targetare import InterrogationResult


def build_workbook(rows):
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    stream = io.BytesIO()
    workbook.save(stream)
    workbook.close()
    stream.seek(0)
    return stream


def make_app(tmp_path):
    database = tmp_path / "test.sqlite3"
    workbook = tmp_path / "active.xlsx"
    return create_app(
        {
            "TESTING": True,
            "DATABASE": str(database),
            "ACTIVE_WORKBOOK": str(workbook),
            "TARGETARE_API_KEY": "secret",
            "SECRET_KEY": "test",
        }
    )


def test_default_pagination_is_100_rows(tmp_path):
    app = make_app(tmp_path)
    assert app.config["COMPANIES_PER_PAGE"] == 100


def test_upload_xlsx_creates_downloadable_workbook_with_contact_columns(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    stream = build_workbook(
        [
            ["Raport firme"],
            ["Denumire", "CUI", "Adresa"],
            ["ACME SRL", "RO12345678", "Bucuresti"],
        ]
    )

    response = client.post(
        "/upload",
        data={"file": (stream, "companies.xlsx")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 200
    assert b"ACME SRL" in response.data
    assert b"Descarc" in response.data

    workbook = load_workbook(app.config["ACTIVE_WORKBOOK"], data_only=True)
    worksheet = workbook.active
    headers = [cell.value for cell in worksheet[2]]
    assert "Emailuri Targetare" in headers
    assert "Telefoane Targetare" in headers
    workbook.close()

    download = client.get("/download")
    assert download.status_code == 200
    assert download.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_interrogate_company_saves_results_to_xlsx(tmp_path, monkeypatch):
    app = make_app(tmp_path)
    client = app.test_client()
    stream = build_workbook(
        [
            ["Denumire", "CUI", "Adresa"],
            ["ACME SRL", "RO12345678", "Bucuresti"],
        ]
    )
    client.post(
        "/upload",
        data={"file": (stream, "companies.xlsx")},
        content_type="multipart/form-data",
    )

    class FakeClient:
        def __init__(self, **_kwargs):
            pass

        def interrogate(self, tax_id):
            assert tax_id == "12345678"
            return InterrogationResult(
                status="success",
                emails={
                    "primaryEmail": "office@example.ro",
                    "secondaryEmail": "sales@example.ro",
                    "websiteEmails": [
                        "office@example.ro",
                        "contact@example.ro",
                    ],
                },
                phones={
                    "primaryPhone": "+40700000000",
                    "secondaryPhone": "+40210000000",
                    "verifiedPhones": ["+40700000000"],
                },
                remaining_requests=98,
            )

    monkeypatch.setattr("targetare_contacts.TargetareClient", FakeClient)

    with app.app_context():
        db = get_db()
        company_id = db.execute("SELECT id FROM companies").fetchone()["id"]

    response = client.post(
        f"/companies/{company_id}/interrogate", follow_redirects=True
    )

    assert response.status_code == 200
    assert b"office@example.ro" in response.data
    assert b"salvat" in response.data

    workbook = load_workbook(app.config["ACTIVE_WORKBOOK"], data_only=True)
    worksheet = workbook.active
    headers = {cell.value: cell.column for cell in worksheet[1]}
    assert worksheet.cell(2, headers["Emailuri Targetare"]).value == (
        "office@example.ro; sales@example.ro; contact@example.ro"
    )
    assert worksheet.cell(2, headers["Telefoane Targetare"]).value == (
        "+40700000000; +40210000000"
    )
    workbook.close()


def test_upload_csv_generates_xlsx_working_copy(tmp_path):
    app = make_app(tmp_path)
    client = app.test_client()
    csv_data = "Denumire;CUI;Adresa\nACME SRL;RO12345678;Bucuresti\n"

    response = client.post(
        "/upload",
        data={"file": (io.BytesIO(csv_data.encode()), "companies.csv")},
        content_type="multipart/form-data",
        follow_redirects=True,
    )

    assert response.status_code == 200
    workbook = load_workbook(app.config["ACTIVE_WORKBOOK"], data_only=True)
    worksheet = workbook.active
    assert worksheet["A2"].value == "ACME SRL"
    assert worksheet["D1"].value == "Emailuri Targetare"
    assert worksheet["E1"].value == "Telefoane Targetare"
    workbook.close()
