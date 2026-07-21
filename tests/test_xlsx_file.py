from openpyxl import Workbook, load_workbook

from targetare_contacts.xlsx_file import write_company_contacts


def test_write_company_contacts_reuses_existing_columns(tmp_path):
    path = tmp_path / "companies.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(
        [
            "Denumire",
            "CUI",
            "Emailuri Targetare",
            "Telefoane Targetare",
            "Status interogare",
        ]
    )
    worksheet.append(
        ["ACME SRL", "12345678", "old@example.ro", "+40111111111", "Interogat"]
    )
    workbook.save(path)
    workbook.close()

    write_company_contacts(
        path,
        source_row=2,
        emails=["new@example.ro", "new@example.ro"],
        phones=["+40700000000", "+40700000000"],
        status="success",
    )

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.active
    assert worksheet["C2"].value == "new@example.ro"
    assert worksheet["D2"].value == "+40700000000"
    assert worksheet["E2"].value == "Interogat"
    assert worksheet.max_column == 5
    workbook.close()
