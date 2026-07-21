from __future__ import annotations

from typing import BinaryIO
from zipfile import BadZipFile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from .csv_import import ALIASES, CompanyRow, ImportReport, _normalize_header, normalize_tax_id


class XLSXImportError(ValueError):
    pass


def _find_column_index(headers: list[str], alias_group: str, required: bool) -> int | None:
    for index, header in enumerate(headers):
        if _normalize_header(header) in ALIASES[alias_group]:
            return index

    if required:
        available = ", ".join(header for header in headers if header)
        raise XLSXImportError(
            f"Nu am găsit coloana pentru {alias_group}. Coloane detectate: {available}."
        )
    return None


def _find_header_row(worksheet) -> tuple[int, list[str]]:
    max_scan_rows = min(worksheet.max_row or 1, 25)

    for row_number, values in enumerate(
        worksheet.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1
    ):
        headers = [str(value).strip() if value is not None else "" for value in values]
        normalized = {_normalize_header(header) for header in headers if header}
        has_name = bool(normalized & ALIASES["company_name"])
        has_tax_id = bool(normalized & ALIASES["tax_id"])
        if has_name and has_tax_id:
            return row_number, headers

    raise XLSXImportError(
        "Nu am găsit un rând de antet care să conțină denumirea firmei și CUI-ul."
    )


def parse_companies_xlsx(stream: BinaryIO) -> tuple[list[CompanyRow], ImportReport]:
    try:
        workbook = load_workbook(stream, read_only=True, data_only=True)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise XLSXImportError("Fișierul XLSX nu este valid sau este corupt.") from exc

    try:
        if not workbook.sheetnames:
            raise XLSXImportError("Fișierul XLSX nu conține nicio foaie de calcul.")

        worksheet = workbook[workbook.sheetnames[0]]
        header_row, headers = _find_header_row(worksheet)
        name_index = _find_column_index(headers, "company_name", required=True)
        tax_id_index = _find_column_index(headers, "tax_id", required=True)
        address_index = _find_column_index(headers, "address", required=False)

        rows: list[CompanyRow] = []
        seen_tax_ids: set[str] = set()
        duplicates = 0
        invalid_tax_ids = 0

        for source_row, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not any(value not in (None, "") for value in values):
                continue

            def value_at(index: int | None):
                if index is None or index >= len(values):
                    return None
                return values[index]

            tax_id = normalize_tax_id(value_at(tax_id_index))
            if not tax_id or len(tax_id) > 10:
                invalid_tax_ids += 1
                continue
            if tax_id in seen_tax_ids:
                duplicates += 1
                continue

            company_name = str(value_at(name_index) or "").strip()
            if not company_name:
                company_name = f"Firmă CUI {tax_id}"

            address = str(value_at(address_index) or "").strip()
            rows.append(
                CompanyRow(
                    company_name=company_name,
                    tax_id=tax_id,
                    original_address=address,
                    source_row=source_row,
                )
            )
            seen_tax_ids.add(tax_id)

        if not rows:
            raise XLSXImportError("Nu am găsit nicio firmă cu un CUI valid.")

        return rows, ImportReport(
            imported=len(rows), duplicates=duplicates, invalid_tax_ids=invalid_tax_ids
        )
    finally:
        workbook.close()
