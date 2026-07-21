from __future__ import annotations

import os
import shutil
from io import BytesIO
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .csv_import import CompanyRow
from .xlsx_import import _find_header_row

EMAIL_HEADER = "Emailuri Targetare"
PHONE_HEADER = "Telefoane Targetare"
STATUS_HEADER = "Status interogare"


class WorkbookUpdateError(RuntimeError):
    pass


def _unique_text(values: Iterable[object]) -> str:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        item = str(value or "").strip()
        key = item.casefold()
        if item and key not in seen:
            result.append(item)
            seen.add(key)
    return "; ".join(result)


def _status_text(status: str) -> str:
    return {
        "success": "Interogat",
        "partial": "Parțial",
        "error": "Eroare",
    }.get(status, "")


def _ensure_tracking_columns(worksheet, header_row: int) -> tuple[int, int, int]:
    email_column = phone_column = status_column = None
    last_header_column = 0
    for cell in worksheet[header_row]:
        value = str(cell.value or "").strip()
        if value:
            last_header_column = max(last_header_column, cell.column)
        normalized = value.casefold()
        if normalized == EMAIL_HEADER.casefold():
            email_column = cell.column
        elif normalized == PHONE_HEADER.casefold():
            phone_column = cell.column
        elif normalized == STATUS_HEADER.casefold():
            status_column = cell.column

    next_column = last_header_column + 1
    if email_column is None:
        email_column = next_column
        worksheet.cell(row=header_row, column=email_column, value=EMAIL_HEADER)
        next_column += 1
    if phone_column is None:
        phone_column = next_column
        worksheet.cell(row=header_row, column=phone_column, value=PHONE_HEADER)
        next_column += 1
    if status_column is None:
        status_column = next_column
        worksheet.cell(row=header_row, column=status_column, value=STATUS_HEADER)
    return email_column, phone_column, status_column


def _save_atomic(workbook, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary = destination.with_name(f".{destination.name}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    except Exception as exc:
        raise WorkbookUpdateError(f"Nu am putut salva fișierul XLSX: {exc}") from exc
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)


def backup_active_workbook(source: str | Path, backup: str | Path) -> bool:
    source_path = Path(source)
    if not source_path.is_file():
        return False
    backup_path = Path(backup)
    backup_path.parent.mkdir(parents=True, exist_ok=True)
    temporary = backup_path.with_name(f".{backup_path.name}.tmp")
    try:
        shutil.copy2(source_path, temporary)
        os.replace(temporary, backup_path)
    except OSError as exc:
        raise WorkbookUpdateError(f"Nu am putut crea copia de siguranță XLSX: {exc}") from exc
    finally:
        if temporary.exists():
            temporary.unlink(missing_ok=True)
    return True


def _write_rows(worksheet, rows: list[CompanyRow], header_row: int) -> None:
    email_column, phone_column, status_column = _ensure_tracking_columns(
        worksheet, header_row
    )
    for row in rows:
        worksheet.cell(
            row=row.source_row,
            column=email_column,
            value=_unique_text(row.imported_emails),
        )
        worksheet.cell(
            row=row.source_row,
            column=phone_column,
            value=_unique_text(row.imported_phones),
        )
        worksheet.cell(
            row=row.source_row,
            column=status_column,
            value=_status_text(row.imported_status),
        )


def prepare_uploaded_xlsx(
    raw: bytes,
    destination: str | Path,
    rows: list[CompanyRow] | None = None,
) -> None:
    try:
        workbook = load_workbook(BytesIO(raw))
    except Exception as exc:
        raise WorkbookUpdateError(
            "Fișierul XLSX nu a putut fi pregătit pentru salvare."
        ) from exc
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_row, _headers = _find_header_row(worksheet)
        if rows is not None:
            _write_rows(worksheet, rows, header_row)
        else:
            email_column, phone_column, status_column = _ensure_tracking_columns(
                worksheet, header_row
            )
            for row_number in range(header_row + 1, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row_number, column=status_column)
                if str(status_cell.value or "").strip():
                    continue
                email_value = worksheet.cell(row=row_number, column=email_column).value
                phone_value = worksheet.cell(row=row_number, column=phone_column).value
                if str(email_value or "").strip() or str(phone_value or "").strip():
                    status_cell.value = "Interogat"
        _save_atomic(workbook, Path(destination))
    finally:
        workbook.close()


def create_xlsx_from_rows(rows: list[CompanyRow], destination: str | Path) -> None:
    workbook = Workbook()
    try:
        worksheet = workbook.active
        worksheet.title = "Companii"
        worksheet.append(["Denumire", "Cod unic inregistrare", "Adresa"])
        for row in rows:
            worksheet.cell(row=row.source_row, column=1, value=row.company_name)
            worksheet.cell(row=row.source_row, column=2, value=row.tax_id)
            worksheet.cell(row=row.source_row, column=3, value=row.original_address)
        _write_rows(worksheet, rows, 1)
        _save_atomic(workbook, Path(destination))
    finally:
        workbook.close()


def write_company_contacts(
    workbook_path: str | Path,
    source_row: int,
    emails: list[str] | None,
    phones: list[str] | None,
    status: str = "success",
) -> None:
    path = Path(workbook_path)
    if not path.is_file():
        raise WorkbookUpdateError("Fișierul XLSX activ nu mai există.")
    try:
        workbook = load_workbook(BytesIO(path.read_bytes()))
    except Exception as exc:
        raise WorkbookUpdateError("Fișierul XLSX activ nu poate fi deschis.") from exc
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        header_row, _headers = _find_header_row(worksheet)
        email_column, phone_column, status_column = _ensure_tracking_columns(
            worksheet, header_row
        )
        if emails is not None:
            worksheet.cell(
                row=source_row,
                column=email_column,
                value=_unique_text(emails),
            )
        if phones is not None:
            worksheet.cell(
                row=source_row,
                column=phone_column,
                value=_unique_text(phones),
            )
        worksheet.cell(
            row=source_row,
            column=status_column,
            value=_status_text(status),
        )
        _save_atomic(workbook, path)
    finally:
        workbook.close()
